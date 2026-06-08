"""Tests for the new upload diagnostic feature:
 - REQUIRED_ODDS now has 11 entries (1, X, 2, U/O 1.5/2.5/3.5, GG, NG); 1X/X2/12 are estimated.
 - parse_excel_bytes returns dict {matches, skipped, rows_seen}.
 - POST /api/upload-excel response includes inserted/updated/unchanged/skipped/total_parsed/rows_seen
   and persists the diagnostic into MongoDB collection upload_skipped (_id='latest').
 - GET /api/upload/skipped returns the last upload diagnostic (or empty defaults when nothing was uploaded).
"""
import io
import pytest
import openpyxl


# ---------- helpers ----------

def _wb_row(time, manif, sq1, sq2, *,
            o1=2.0, oX=3.2, o2=3.5, o1X=None, oX2=None, o12=None,
            oU15=1.40, oO15=2.80, oU25=1.85, oO25=1.95,
            oU35=1.30, oO35=3.40, oGG=1.85, oNG=1.95):
    """Build an Excel row with the layout expected by the parser.
    Columns (1-based): A=time, B=manif, E=sq1, F=sq2, I/J/K=1/X/2,
    P/Q/R=1X/X2/12, S/T=U1.5/O1.5, U/V=U2.5/O2.5, W/X=U3.5/O3.5, Y/Z=GG/NG."""
    return [time, manif, "S", "X", sq1, sq2, "", "",
            o1, oX, o2, -1, 1.4, 2.5, 1.8,
            o1X, oX2, o12,
            oU15, oO15, oU25, oO25, oU35, oO35, oGG, oNG]


def _build_excel(rows, header="matches, 18 maggio") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = header
    for r, row in enumerate(rows, start=5):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c).value = v
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@pytest.fixture(autouse=True)
def _clean(api_client, base_url):
    api_client.delete(f"{base_url}/api/matches/all", timeout=15)
    # Also wipe the persisted diagnostic so tests are deterministic
    api_client.delete(f"{base_url}/api/matches/all", timeout=15)
    yield


def _wipe_skipped_doc(api_client, base_url):
    """Direct mongo doesn't have a route, but we can simulate empty by replacing
    the file the diagnostic stores; easiest is to just call upload with an
    empty workbook. However the cleanest is via a tiny upload that resets it
    to the new state. For testing the empty case we run BEFORE any upload."""
    pass


# ---------- TESTS ----------

class TestUploadSkippedDiagnostic:

    def test_skipped_endpoint_clean_state(self, api_client, base_url):
        """When NO upload has ever been done, GET /api/upload/skipped returns
        a dict with null filename / null uploaded_at / 0 counters / empty list.
        We force a clean state by deleting the upload_skipped doc via direct mongo.
        """
        # Use motor-equivalent via PyMongo (sync) to wipe the doc deterministically
        import pymongo, os
        from dotenv import load_dotenv
        from pathlib import Path
        load_dotenv(Path('/app/backend/.env'))
        client = pymongo.MongoClient(os.environ['MONGO_URL'])
        db = client[os.environ['DB_NAME']]
        db.upload_skipped.delete_many({})
        client.close()

        r = api_client.get(f"{base_url}/api/upload/skipped", timeout=10)
        assert r.status_code == 200
        data = r.json()
        assert data == {
            "filename": None,
            "uploaded_at": None,
            "rows_seen": 0,
            "valid_matches": 0,
            "inserted": 0,
            "updated": 0,
            "unchanged": 0,
            "skipped_count": 0,
            "skipped": [],
        }

    def test_upload_response_shape(self, api_client, base_url):
        """A valid upload must return all 6 expected keys."""
        rows = [_wb_row("12:00", "TEST_LIGA", "TEST_HomeA", "TEST_AwayA")]
        content = _build_excel(rows)
        files = {'file': ('valid.xlsx', content,
                          'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        r = api_client.post(f"{base_url}/api/upload-excel", files=files, timeout=30)
        assert r.status_code == 200, r.text
        data = r.json()
        for key in ("inserted", "updated", "unchanged", "skipped",
                    "total_parsed", "rows_seen"):
            assert key in data, f"missing key {key} in {data}"
        assert data["inserted"] == 1
        assert data["total_parsed"] == 1
        assert data["skipped"] == 0
        assert data["rows_seen"] >= 1

    def test_doubles_chance_only_missing_is_valid(self, api_client, base_url):
        """A match missing ONLY 1X/X2/12 must be VALID (those are estimated).
        REQUIRED_ODDS has 11 quotes — none of them is 1X/X2/12."""
        rows = [_wb_row("13:00", "TEST_DC", "TEST_DCHome", "TEST_DCAway",
                        o1X=None, oX2=None, o12=None)]
        content = _build_excel(rows)
        files = {'file': ('dc.xlsx', content,
                          'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        r = api_client.post(f"{base_url}/api/upload-excel", files=files, timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert data["inserted"] == 1, f"got {data}"
        assert data["skipped"] == 0

        # Verify the match has odd_1X/X2/12 filled (estimated)
        r2 = api_client.get(f"{base_url}/api/matches", params={"q": "DCHome"}, timeout=10)
        ms = r2.json()
        assert len(ms) == 1
        odds = ms[0]["odds"]
        for fld in ("odd_1X", "odd_X2", "odd_12"):
            assert odds.get(fld) is not None, f"{fld} not estimated"
            assert fld in odds.get("estimated", []), f"{fld} missing from estimated list"

    def test_missing_required_O25_is_skipped_with_reason(self, api_client, base_url):
        """A row missing odd_O25 (REQUIRED) must be SKIPPED and appear in the
        diagnostic with reason citing the missing field + odds_read populated."""
        bad_row = _wb_row("14:00", "TEST_BAD", "TEST_BadHome", "TEST_BadAway",
                          oO25=None)
        good_row = _wb_row("15:00", "TEST_GOOD", "TEST_GoodHome", "TEST_GoodAway")
        content = _build_excel([bad_row, good_row])
        files = {'file': ('miss.xlsx', content,
                          'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        r = api_client.post(f"{base_url}/api/upload-excel", files=files, timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert data["inserted"] == 1
        assert data["skipped"] == 1
        assert data["total_parsed"] == 1
        assert data["rows_seen"] == 2

        # Diagnostic must report the skipped row
        r2 = api_client.get(f"{base_url}/api/upload/skipped", timeout=10)
        assert r2.status_code == 200
        diag = r2.json()
        assert diag["filename"] == "miss.xlsx"
        assert diag["uploaded_at"] is not None
        assert diag["skipped_count"] == 1
        assert len(diag["skipped"]) == 1
        sk = diag["skipped"][0]
        assert sk["sq1"] == "TEST_BadHome"
        assert sk["sq2"] == "TEST_BadAway"
        assert "odd_O25" in sk["reason"], f"reason should cite missing odd_O25: {sk['reason']}"
        assert "odd_O25" in sk["missing"], f"missing list should contain odd_O25: {sk['missing']}"
        # odds_read must contain the quotes that WERE read (e.g. odd_1, odd_GG)
        assert "odd_1" in sk["odds_read"]
        assert "odd_GG" in sk["odds_read"]
        # Must NOT contain odd_O25 since it was missing
        assert "odd_O25" not in sk["odds_read"]

    def test_missing_teams_skipped_with_reason(self, api_client, base_url):
        """A row with empty squadra1/squadra2 must be skipped with reason 'Squadre mancanti'."""
        no_team_row = _wb_row("16:00", "TEST_NOTEAM", "", "")
        good_row = _wb_row("17:00", "TEST_GOOD2", "TEST_G2Home", "TEST_G2Away")
        content = _build_excel([no_team_row, good_row])
        files = {'file': ('noteam.xlsx', content,
                          'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        r = api_client.post(f"{base_url}/api/upload-excel", files=files, timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert data["inserted"] == 1
        assert data["skipped"] == 1

        r2 = api_client.get(f"{base_url}/api/upload/skipped", timeout=10)
        diag = r2.json()
        assert diag["skipped_count"] == 1
        sk = diag["skipped"][0]
        assert sk["reason"] == "Squadre mancanti", f"unexpected reason: {sk['reason']}"

    def test_required_odds_list_size_via_skipped(self, api_client, base_url):
        """Upload a row missing ALL 11 required quotes — the 'missing' list
        in the diagnostic must contain exactly the 11 REQUIRED_ODDS items."""
        empty_row = _wb_row("18:00", "TEST_ALLMISS", "TEST_AllMissHome", "TEST_AllMissAway",
                            o1=None, oX=None, o2=None,
                            oU15=None, oO15=None, oU25=None, oO25=None,
                            oU35=None, oO35=None, oGG=None, oNG=None)
        content = _build_excel([empty_row])
        files = {'file': ('allmiss.xlsx', content,
                          'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        r = api_client.post(f"{base_url}/api/upload-excel", files=files, timeout=30)
        assert r.status_code == 200
        assert r.json()["skipped"] == 1

        diag = api_client.get(f"{base_url}/api/upload/skipped", timeout=10).json()
        assert diag["skipped_count"] == 1
        missing = diag["skipped"][0]["missing"]
        expected = {'odd_1', 'odd_X', 'odd_2',
                    'odd_U15', 'odd_O15', 'odd_U25', 'odd_O25',
                    'odd_U35', 'odd_O35', 'odd_GG', 'odd_NG'}
        assert set(missing) == expected, f"REQUIRED_ODDS should be exactly 11; got {missing}"
        assert len(missing) == 11


class TestRegressionEndpoints:
    """Regression: ensure existing endpoints still work after the refactor."""

    def test_matches_days_endpoint(self, api_client, base_url):
        r = api_client.get(f"{base_url}/api/matches/days", timeout=10)
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_matches_filter_by_day(self, api_client, base_url):
        # seed one match
        rows = [_wb_row("19:00", "TEST_REG", "TEST_RegHome", "TEST_RegAway")]
        files = {'file': ('reg.xlsx', _build_excel(rows),
                          'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        api_client.post(f"{base_url}/api/upload-excel", files=files, timeout=30)

        days = api_client.get(f"{base_url}/api/matches/days", timeout=10).json()
        assert len(days) >= 1
        r = api_client.get(f"{base_url}/api/matches", params={"day": days[0]}, timeout=10)
        assert r.status_code == 200
        ms = r.json()
        assert isinstance(ms, list)
        for m in ms:
            assert m["day"] == days[0]
            assert "_id" not in m

    def test_matches_selected_list(self, api_client, base_url):
        r = api_client.get(f"{base_url}/api/matches/selected/list", timeout=10)
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_ml_stats(self, api_client, base_url):
        r = api_client.get(f"{base_url}/api/ml/stats", timeout=10)
        assert r.status_code == 200
        data = r.json()
        assert "markets" in data
        assert "family_totals" in data
        assert isinstance(data["markets"], list)
