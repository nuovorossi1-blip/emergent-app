import os
import io
import pytest
import requests
from pathlib import Path
import openpyxl

# Load EXPO_BACKEND_URL from /app/frontend/.env (Expo uses EXPO_PUBLIC_BACKEND_URL)
def _load_backend_url() -> str:
    fe_env = Path('/app/frontend/.env')
    if fe_env.exists():
        for line in fe_env.read_text().splitlines():
            if line.startswith('EXPO_PUBLIC_BACKEND_URL='):
                return line.split('=', 1)[1].strip().strip('"')
    raise RuntimeError("EXPO_PUBLIC_BACKEND_URL not found in /app/frontend/.env")


BASE_URL = _load_backend_url().rstrip('/')


@pytest.fixture(scope='session')
def base_url():
    return BASE_URL


@pytest.fixture(scope='session')
def api_client():
    s = requests.Session()
    return s


def _make_excel_full() -> bytes:
    """6 matches across 2 days (day rollover at row 5 when time goes back from 20:45 to 10:00)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "matches, 12 maggio"
    data = [
        ["11:30", "AUS SEG", "S", "X", "TEST_Canberra White", "TEST_Queanbeyan City Fc", "", "", 7.50, 5.75, 1.24, -1, 3.40, 4.25, 1.72, 3.00, 1.04, 1.10, 8.00, 1.02, 4.00, 1.18, 2.20, 1.52, 1.40, 2.75],
        ["14:00", "BRA2", "S", "X", "TEST_Botafogo Sp", "TEST_Athletic Club", "", "", 2.00, 3.10, 3.75, -1, 1.40, 2.50, 1.85, 1.22, 1.67, 1.30, 2.80, 1.36, 1.57, 2.20, 1.16, 4.00, 1.95, 1.72],
        ["18:00", "SERIE A", "S", "X", "TEST_Inter Milan", "TEST_Roma", "", "", 1.80, 3.60, 4.40, -1, 1.30, 2.40, 2.10, 1.20, 1.70, 1.35, 3.50, 1.25, 1.85, 1.95, 1.30, 3.50, 1.75, 1.95],
        ["20:45", "SERIE A", "S", "X", "TEST_Juventus", "TEST_Lazio", "", "", 1.55, 4.00, 6.00, -1, 1.20, 2.80, 2.40, 1.15, 1.85, 1.42, 4.20, 1.20, 2.10, 1.75, 1.40, 2.85, 1.85, 1.85],
        # Day rollover triggered here (10:00 < 20:45)
        ["10:00", "PREMIER", "S", "X", "TEST_Manchester Utd", "TEST_Arsenal", "", "", 2.40, 3.30, 2.85, -1, 1.50, 2.60, 2.00, 1.30, 1.70, 1.45, 3.10, 1.30, 1.75, 2.10, 1.20, 3.80, 1.80, 1.90],
        ["15:30", "BUNDES", "S", "X", "TEST_Bayern Munich", "TEST_Dortmund", "", "", 1.70, 3.80, 4.50, -1, 1.25, 2.50, 2.30, 1.18, 1.85, 1.40, 4.00, 1.22, 1.90, 1.90, 1.32, 3.30, 1.70, 2.00],
    ]
    for r, row in enumerate(data, start=5):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c).value = v
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _make_excel_partial() -> bytes:
    """One match where ONLY 1X, X2, 12 (doubles chance) are missing.
    Under the new REQUIRED_ODDS (11 quotes), doubles chance are NOT required
    and must be estimated. All other quotes (1/X/2, U/O 1.5/2.5/3.5, GG/NG)
    are present so the match must be VALID."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "matches, 15 maggio"
    row = ["12:00", "TESTPART", "S", "X", "TEST_PartialA", "TEST_PartialB", "", "",
           2.0,   # 1
           3.2,   # X
           3.5,   # 2
           -1, 1.4, 2.5, 1.8,
           None,  # 1X missing (not required, will be estimated)
           None,  # X2 missing (not required, will be estimated)
           None,  # 12 missing (not required, will be estimated)
           1.6,   # U15
           2.3,   # O15
           1.5,   # U25
           2.4,   # O25
           1.2,   # U35
           3.8,   # O35
           1.9,   # GG
           1.8]   # NG
    for c, v in enumerate(row, start=1):
        ws.cell(row=5, column=c).value = v
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _make_excel_missing_required() -> bytes:
    """Two rows: one missing odd_O25 (required) -> should be skipped.
    One complete row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "matches, 20 maggio"
    row_bad = ["13:00", "MISSREQ", "S", "X", "TEST_BadA", "TEST_BadB", "", "",
               2.0, 3.2, 3.5, -1, 1.4, 2.5, 1.8, 1.3, 1.5, 1.4, 1.5, 2.5, 1.6, None, 1.3, 3.2, 1.9, 1.8]  # O25=None
    row_ok = ["16:00", "OKREQ", "S", "X", "TEST_OkA", "TEST_OkB", "", "",
              2.0, 3.2, 3.5, -1, 1.4, 2.5, 1.8, 1.3, 1.5, 1.4, 1.5, 2.5, 1.6, 2.4, 1.3, 3.2, 1.9, 1.8]
    for c, v in enumerate(row_bad, start=1):
        ws.cell(row=5, column=c).value = v
    for c, v in enumerate(row_ok, start=1):
        ws.cell(row=6, column=c).value = v
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@pytest.fixture(scope='session')
def excel_full() -> bytes:
    return _make_excel_full()


@pytest.fixture(scope='session')
def excel_partial() -> bytes:
    return _make_excel_partial()


@pytest.fixture(scope='session')
def excel_missing_required() -> bytes:
    return _make_excel_missing_required()
